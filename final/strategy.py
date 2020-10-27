import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment


class Strategy:
    def __init__(self, files, ocurr=0, hr=0, minute=0):#revisar construtor
        self.files = files
        self.ocurr = ocurr
        self.hr = hr
        self.minute = minute
        self.listequip = []
        self.listocurr = []
        self.listtime = []
        self.listporc = []

    def formatcontent(self, *args):
        args[0].create_sheet(args[1])
        ws = args[0].get_sheet_by_name(args[1])
        df = pd.DataFrame(args[2], columns=['timestamp', 'COMS STATUS'])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        datas = self.recorddata(ws, args[1], df)
        self.formatsheet(ws)
        return datas

    @staticmethod
    def formatsheet(ws):
        flags = ['Equipamento:', 'Eventos:', 'Tempo total:', 'Indisp no mês:']
        for line, notification in enumerate(flags):
            ws.cell(row=line+2, column=5).value = notification
            ws.cell(row=line+2, column=5).font = Font(bold=True)
            ws.cell(row=line+2, column=6).alignment = Alignment(horizontal='left')

    def recorddata(self, *args):
        self.ocurr = args[2].shape[0]
        sheet_name = '{}'.format(args[1])
        calctime = self.calctime(self.ocurr)
        porcen = '{}%'.format(round(self.calcporc(), 2))
        generaldatas = [sheet_name, self.ocurr, calctime, porcen]
        for line, gen in enumerate(generaldatas):
            args[0].cell(row=line + 2, column=6).value = gen
        return sheet_name, self.ocurr, calctime, porcen

    def listdata(self, *args):
        self.listequip.append(args[0])
        self.listocurr.append(args[1])
        self.listtime.append(args[2])
        self.listporc.append(args[3])
        return self.listequip, self.listocurr, self.listtime, self.listporc

    def calctime(self, noc):
        self.hr = (noc // 4)
        self.minute = ((noc - (self.hr * 4)) * 15)
        return '{} hs : {} min : {} seg'.format(self.hr, self.minute, 0)

    def calcporc(self):
        tothr = (31 * 24)
        convhrs = self.minute / 60
        result = ((self.hr + convhrs) * 100) / tothr
        return result

    @staticmethod
    def report(*args):
        wt = args[0].get_sheet_by_name('Sheet')
        wt.title = 'Relatório-Steag'
        obj = {'Equipamento': args[1], 'Eventos': args[2], 'Tempo': args[3], '% indisp': args[4]}
        df_report = pd.DataFrame(data=obj)
        df_report.sort_values('Eventos', ascending=False, inplace=True)
        for r in dataframe_to_rows(df_report, index=False, header=True):
            wt.append(r)
