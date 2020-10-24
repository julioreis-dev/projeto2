import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment


class InverterFilter:
    def __init__(self, files, ocurr=0, hr=0, minute=0, listequip=None, listocurr=None, listtime=None, listporc=None):
        if listporc is None:
            listporc = []
        if listtime is None:
            listtime = []
        if listocurr is None:
            listocurr = []
        if listequip is None:
            listequip = []

        self.files = files
        self.ocurr = ocurr
        self.hr = hr
        self.minute = minute
        self.listequip = listequip
        self.listocurr = listocurr
        self.listtime = listtime
        self.listporc = listporc

    def filter(self):
        wb = Workbook()
        number = 0
        for n in self.files:
            label = n[-22:-4]
            df = pd.read_csv(n)
            df_read = df.loc[df['COMS STATUS'] == 255]
            del df_read['ACTIVE POWER']
            if not df_read.empty:
                general_datas = self.formatcontent(wb, label, df_read)
                dataslist = self.listdata(general_datas[0], general_datas[1], general_datas[2], general_datas[3])
                number += 1
        self.report(wb, dataslist[0], dataslist[1], dataslist[2], dataslist[3])
        wb.save('teste1.xlsx')
        wb.close()
        print('{} Inversores possuem registro de indisponibilidade.'.format(number))

    def formatcontent(self, wb, label, df_read):
        wb.create_sheet(label)
        ws = wb.get_sheet_by_name(label)
        df_read = pd.DataFrame(df_read, columns=['timestamp', 'COMS STATUS'])
        for r in dataframe_to_rows(df_read, index=False, header=True):
            ws.append(r)
        datas = self.recorddata(ws, label, df_read)
        self.formatsheet(ws)
        return datas

    @staticmethod
    def formatsheet(ws):
        ws['e2'] = 'Equipamento:'
        ws['e3'] = 'Ocorrencias:'
        ws['e4'] = 'Tempo total:'
        ws['e5'] = 'Indisp no mês:'
        ws['f2'].alignment = Alignment(horizontal='left')
        ws['f3'].alignment = Alignment(horizontal='left')
        ws['f4'].alignment = Alignment(horizontal='left')
        ws['f5'].alignment = Alignment(horizontal='left')

    def recorddata(self, ws, label, df_read):
        self.ocurr = df_read.shape[0]
        sheet_name = label
        ocurr = self.ocurr
        calctime = self.calctime()
        porcen = round(self.calcporc(), 2)
        ws['f2'] = '{}'.format(sheet_name)
        ws['f3'] = ocurr
        ws['f4'] = calctime
        ws['f5'] = '{}%'.format(porcen)
        return sheet_name, ocurr, calctime, porcen

    def listdata(self, *args):
        self.listequip.append(args[0])
        self.listocurr.append(args[1])
        self.listtime.append(args[2])
        self.listporc.append(args[3])
        return self.listequip, self.listocurr, self.listtime, self.listporc

    def calctime(self):
        self.hr = (self.ocurr // 4)
        self.minute = (self.ocurr - (self.hr * 4)) * 15
        return '{} hs : {} min : {} seg'.format(self.hr, self.minute, 0)

    def calcporc(self):
        tothr = (31 * 24)
        convhrs = self.minute / 60
        result = ((self.hr + convhrs) * 100) / tothr
        return result

    @staticmethod
    def report(*args):
        wt = args[0].get_sheet_by_name('Sheet')
        obj = {'Equipamento': args[1], 'Ocorrências': args[2], 'Tempo': args[3], '% indisp': args[4]}
        df_report = pd.DataFrame(data=obj)
        for r in dataframe_to_rows(df_report, index=False, header=True):
            wt.append(r)
