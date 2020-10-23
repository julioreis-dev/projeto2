import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class BoxFilter:
    def __init__(self, files, ocurr=0, hr=0, minute=0):
        self.files = files
        self.ocurr = ocurr
        self.hr = hr
        self.minute = minute

    def filter(self):
        wb = Workbook()
        number = 0
        for n in self.files:
            label = n[-28:-4]
            df = pd.read_csv(n)
            df_read = df.loc[df['COMS STATUS'] == 255]
            del df_read['Current']
            del df_read['Power']
            if not df_read.empty:
                self.formatcontent(wb, label, df_read)
                number += 1
        print('{} strings possuem dados de indisponibilidade.'.format(number))
        wb.save('teste2.xlsx')

    def formatcontent(self, wb, label, df_read):
        wb.create_sheet(label)
        ws = wb.get_sheet_by_name(label)
        df_read = pd.DataFrame(df_read, columns=['EQUIPAMENTO', 'timestamp', 'COMS STATUS'])
        df_read = df_read.fillna(label)
        for r in dataframe_to_rows(df_read, index=False, header=True):
            ws.append(r)
        self.ocurr = df_read.shape[0]
        ws['f2'] = self.ocurr
        ws['f3'] = self.calctime()
        ws['f4'] = 'O equipamento "{}" ficou {}% indisponivel no mês'.format(label, (round(self.calcporc(), 2)))

    def calctime(self):
        self.hr = (self.ocurr // 4)
        self.minute = (self.ocurr - (self.hr * 4)) * 15
        return 'Tempo de indisponibilidade: {} hs : {} min : {} seg'.format(self.hr, self.minute, 0)

    def calcporc(self):
        tothr = (31 * 24)
        convhrs = self.minute/60
        result = ((self.hr+convhrs) * 100) / tothr
        return result
